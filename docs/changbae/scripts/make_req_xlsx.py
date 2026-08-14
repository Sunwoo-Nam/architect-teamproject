# docs/changbae/52·53·54 마크다운 표를 파싱해 요구사항 엑셀 생성
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import os
DOCS = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

def read(p):
    return open(f"{DOCS}/{p}", encoding="utf-8").read()

def cells(line):
    return [c.strip() for c in line.strip().strip("|").split("|")]

def strip_md(s):
    s = re.sub(r"\[([^\]]*)\]\([^)]*\)", r"\1", s)          # 링크 → 텍스트
    s = s.replace("**", "").replace("`", "").replace("<br/>", "\n").replace("\\*", "*")
    s = s.replace("\\(", "(").replace("\\)", ")")
    s = re.sub(r"\*([^*]+)\*", r"\1", s)                     # italic
    s = s.replace("U(r)/U(x\\*)", "U(r)/U(x*)")
    return s.strip()

# ── 52: VoC ──
voc_rows = []  # (인덱스, Stakeholder, VoC, 출처, 분류)
cur_sh = None
for line in read("52-Stakeholder-재편-VoC.md").splitlines():
    m = re.match(r"^### (\w+)\. (.+)$", line)
    if m:
        cur_sh = m.group(2).strip()
        continue
    if re.match(r"^\| [A-Z]{2}-V\d+ ", line):
        c = cells(line)
        if len(c) >= 4:
            voc_rows.append((strip_md(c[0]), cur_sh, strip_md(c[1]).strip('"“”'), strip_md(c[2]), strip_md(c[3])))

# ── 53: FR ──
fr_rows = []  # (ID, 단계, FR, VoC 매핑)
STAGES = {"개시", "선호", "수행", "관찰·개입", "종결", "신뢰·안전", "검증"}
fr_doc = read("53-FR-재도출.md")
fr_main = fr_doc.split("## 53.1")[1].split("### 53.1.1")[0]
for line in fr_main.splitlines():
    if re.match(r"^\| FR-\d+ ", line):
        c = cells(line)
        if len(c) >= 4 and strip_md(c[1]) in STAGES:
            fr_rows.append((strip_md(c[0]), strip_md(c[1]), strip_md(c[2]), strip_md(c[3])))

# ── 54: QAS (본표 + 계층 평가) ──
doc54 = read("54-QAS-재도출.md")
qas_main = {}  # id -> (iso, 시나리오, 측정, voc)
for line in doc54.splitlines():
    if re.match(r"^\| QAS-\d+ \|", line):
        c = cells(line)
        if len(c) >= 5 and "—" in c[1] or (len(c) >= 5 and "Functional" in c[1] or "Safety" in c[1] or "Security" in c[1] or "Performance" in c[1] or "Flexibility" in c[1] or "Reliability" in c[1] or "Interaction" in c[1] or "Maintainability" in c[1]):
            qid = strip_md(c[0])
            if qid not in qas_main and len(c) >= 5:
                qas_main[qid] = (strip_md(c[1]), strip_md(c[2]), strip_md(c[3]), strip_md(c[4]))

# 계층·서열·등급 (§54.2)
tier = {}   # id -> (계층, 서열, 중요도, 난이도, 난이도 근거)
sec = doc54.split("## 54.2")[1].split("## 54.3")[0]
cur_tier = None
for line in sec.splitlines():
    if line.startswith("### 핵심"):
        cur_tier = "핵심"
    elif line.startswith("### 준핵심"):
        cur_tier = "준핵심"
    elif line.startswith("### 나머지"):
        cur_tier = "나머지"
    m = re.match(r"^\| (.+?) \| (QAS-\d+) \| (.+?) \| (.+?) \| (.+?) \|", line)
    if m and cur_tier == "핵심":
        rank, qid, _, grade, basis = m.groups()
        imp, dif = strip_md(grade).split("/")
        tier[qid] = (cur_tier, strip_md(rank), imp, dif, strip_md(basis))
        continue
    m2 = re.match(r"^\| (QAS-\d+) \| (.+?) \| (.+?) \| (.+?) \|", line)
    if m2 and cur_tier in ("준핵심", "나머지"):
        qid, _, grade, basis = m2.groups()
        g = strip_md(grade)
        if "/" in g:
            imp, dif = g.split("/")
        else:
            imp, dif = g, g   # "—" (QAS-02 흡수)
        tier[qid] = (cur_tier, "-", imp, dif, strip_md(basis))

# ═══════ 엑셀 작성 ═══════
KF = "Malgun Gothic"
HD_FILL = PatternFill("solid", fgColor="2A56A5")
HD_FONT = Font(name=KF, size=10, bold=True, color="FFFFFF")
BODY = Font(name=KF, size=10)
BOLD = Font(name=KF, size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CTR = Alignment(horizontal="center", vertical="top", wrap_text=True)
thin = Side(style="thin", color="C9C9C9")
BORD = Border(left=thin, right=thin, top=thin, bottom=thin)
TIER_FILL = {"핵심": PatternFill("solid", fgColor="DAE8FC"),
             "준핵심": PatternFill("solid", fgColor="D5E8D4"),
             "보류": PatternFill("solid", fgColor="F2F2F2")}

wb = Workbook()

def style_sheet(ws, widths, n_rows, center_cols=()):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.fill = HD_FILL
        cell.font = HD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORD
    for r in range(2, n_rows + 2):
        for i in range(1, len(widths) + 1):
            c = ws.cell(row=r, column=i)
            c.font = BODY
            c.alignment = CTR if i in center_cols else WRAP
            c.border = BORD
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(widths))}{n_rows + 1}"

# ── VoC 시트 ──
ws = wb.active
ws.title = "VoC"
ws.append(["인덱스", "Stakeholder", "VoC (1인칭 발언)", "출처", "분류"])
for r in voc_rows:
    ws.append(list(r))
style_sheet(ws, [10, 22, 72, 30, 14], len(voc_rows), center_cols=(1, 5))
for row in ws.iter_rows(min_row=2):
    if "보류" in str(row[4].value):
        for c in row:
            c.fill = TIER_FILL["보류"]

# ── FR 시트 ──
ws = wb.create_sheet("FR")
ws.append(["ID", "단계", "Functional Requirement", "VoC 매핑"])
for r in fr_rows:
    ws.append(list(r))
style_sheet(ws, [9, 11, 86, 22], len(fr_rows), center_cols=(1, 2))

# ── QAS 시트 ──
ws = wb.create_sheet("QAS")
ws.append(["ID", "계층", "서열", "ISO 25010 특성—부특성", "Quality Attribute Scenario", "응답 측정", "중요도", "난이도", "난이도 근거", "VoC 매핑"])
order = sorted(qas_main.keys(), key=lambda k: int(k.split("-")[1]))
for qid in order:
    iso, scen, meas, voc = qas_main[qid]
    t = tier.get(qid, ("나머지", "-", "L", "-", ""))
    ws.append([qid, t[0], t[1], iso, scen, meas, t[2], t[3], t[4], voc])
style_sheet(ws, [9, 9, 6, 26, 56, 26, 8, 8, 44, 16], len(order), center_cols=(1, 2, 3, 7, 8))
for row in ws.iter_rows(min_row=2):
    t = str(row[1].value)
    if t in TIER_FILL:
        for c in row:
            c.fill = TIER_FILL[t]

# ── 개요 시트 ──
ws = wb.create_sheet("개요", 0)
rows = [
    ["다자 협상 요구사항 — VoC · FR · QAS", ""],
    ["", ""],
    ["항목", "건수"],
    ["VoC (Stakeholder 10인)", len(voc_rows)],
    ["FR (협상 라이프사이클)", len(fr_rows)],
    ["QAS (ISO/IEC 25010:2023 매핑)", len(order)],
    ["", ""],
    ["정본 문서", ""],
    ["VoC", "docs/changbae/52-Stakeholder-재편-VoC.md"],
    ["FR", "docs/changbae/53-FR-재도출.md"],
    ["QAS", "docs/changbae/54-QAS-재도출.md"],
    ["", ""],
    ["비고", "본 파일은 정본(md)에서 생성한 사본 (2026-08-14). 건수는 생성 시점 자동 산출값(고정). 정본 변경 시 재생성 필요."],
    ["", "QAS 계층: 핵심 4(파랑) · 준핵심 3(초록) · 나머지 — 중요도는 계층 고정(H/M/L), QAS-02는 QAS-01에 흡수 측정."],
]
for r in rows:
    ws.append(r)
ws.column_dimensions["A"].width = 34
ws.column_dimensions["B"].width = 78
for row in ws.iter_rows():
    for c in row:
        c.font = BODY
        c.alignment = WRAP
ws["A1"].font = Font(name=KF, size=14, bold=True)
ws["A3"].font = BOLD
ws["B3"].font = BOLD
ws["A8"].font = BOLD

out = os.path.join(DOCS, "다자협상-VoC-FR-QAS.xlsx")
wb.save(out)
print(f"VoC {len(voc_rows)} / FR {len(fr_rows)} / QAS {len(order)}")
print("tiers:", {k: tier[k][0] for k in sorted(tier, key=lambda x: int(x.split('-')[1]))})
